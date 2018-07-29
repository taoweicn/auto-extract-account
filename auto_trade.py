import requests
import re
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook


def purchase():

    url = 'http://t.54rj.com/result/'
    headers = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': 'Mozilla/5.0',
    }

    # 提取交易码
    try:
        code = open('data/code.txt', 'r+')
    except FileNotFoundError:
        code = open('data/code.txt', 'w+')
    bills = [item for item in code.read().strip().split() if '201802' in item]
    # 清空code文件
    code.seek(0)
    code.truncate()
    code.close()

    if bills:
        final_result = time.asctime(time.localtime(time.time())) + ' 购入账户\n'
    else:
        print('未找到交易码！')
        return

    available_account = open('data/available_account.txt', 'a+')

    # 打开excel文件
    excel_filename = 'data/purchase_records.xlsx'
    try:
        wb = load_workbook(excel_filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.create_sheet('失效交易码')
        wb.worksheets[0].title = '交易记录'
        wb.worksheets[1].append(['失效交易码'])
        # 设置列宽
        wb.worksheets[0].column_dimensions['A'].width = 25
        wb.worksheets[0].column_dimensions['B'].width = 30
        wb.worksheets[1].column_dimensions['A'].width = 25
    record = wb.worksheets[0]
    record.append([time.asctime(time.localtime(time.time())), '购入账户'])

    count = 0
    for bill in bills:
        result = requests.post(url, headers=headers, data='bill=' + str(bill))
        soup = BeautifulSoup(result.text, 'html.parser')
        reg_str = r'([\w-]+(\.[\w-]+)*@[\w-]+(\.[\w-]+)+)'
        mod = re.compile(reg_str)
        items = mod.findall(str(soup.textarea))
        if not items:
            print('第' + str(bills.index(bill) + 1) + '个未提取到！')
            record.append(['交易码 ' + bill, '未提取到'])
            wb.worksheets[1].append([bill])
        else:
            final_result += '交易码：' + bill + '\n'
            record.append(['交易码：', bill])
            for item in items:
                # print('账户', str(count), ': ', item[0][:-10])
                count += 1
                final_result += '账户' + str(count) + ': ' + item[0][:-10] + ' 密码： ' + item[0][-6:] + '\n'
                record.append(['账户' + str(count), item[0][:-10], 1, '密码', item[0][-6:]])
                available_account.write(item[0][:-10] + '\n')

    # 存入交易记录文件
    wb.save(filename=excel_filename)  # excel记录
    purchase_records = open('data/purchase_records.txt', 'a+')  # 文本记录
    purchase_records.write(final_result)
    purchase_records.close()

    available_account.seek(0)
    print('进货' + str(count) + '个，目前库存为：' + str(len(available_account.read().split())) + '\n')
    # 关闭文件读写
    available_account.close()


def delivery(is_file):
    # 出货前先进货
    purchase()

    available_account = open('data/available_account.txt', 'r+')
    available_account_list = available_account.read().split()
    if not len(available_account_list):
        print('当前库存为0，无法出货')
        available_account.close()
        return

    num = int(input('输入要出货的数量：'))
    if num <= 0:
        return

    if num > len(available_account_list):
        print('库存不够，目前库存：' + str(len(available_account_list)))
        available_account.close()
        return

    # 结果写入文件
    if is_file:
        goods = open('账号' + str(num) + '个.txt', 'w')
        goods.write('密码统一为： zxc009\n')
    print('\n密码统一为： zxc009')
    for i in range(0, num):
        print('账户' + str(i + 1) + ': ' + available_account_list[i])
        if is_file:
            goods.write('账户' + str(i + 1) + ': ' + available_account_list[i] + '\n')

    # 删除原来的数据
    available_account.seek(0)
    available_account.truncate()

    # 写入剩下的数据
    for j in range(num, len(available_account_list)):
        available_account.write(available_account_list[j] + '\n')
    available_account.close()
    if is_file:
        goods.close()
    return


def show_store():
    try:
        available_account = open('data/available_account.txt', 'r')
    except FileNotFoundError:
        available_account = open('data/available_account.txt', 'w+')
    print('\n当前库存：' + str(len(available_account.read().split())))
    return


def main():
    while True:
        show_store()
        choice = input('输入操作： 1出货， 2出货到文件, 0退出  ')
        if choice == '1':
            delivery(is_file=False)
        elif choice == '2':
            delivery(is_file=True)
        elif choice == '0':
            break
        else:
            print('输入有误，请重新输入！')


main()
